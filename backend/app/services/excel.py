"""Excel (.xlsx) rapor üretimi için ortak yardımcı - stok ve fatura
raporları aynı biçimlendirmeyi (kalın başlık, otomatik sütun genişliği)
paylaşsın diye."""

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

MAX_COLUMN_WIDTH = 40


def build_workbook(sheet_title: str, columns: list[str], rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    for i, column_title in enumerate(columns, start=1):
        cell_lengths = [len(str(row[i - 1])) for row in ws.iter_rows(min_row=2, values_only=True)]
        max_len = max([len(column_title), *cell_lengths])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, MAX_COLUMN_WIDTH)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
