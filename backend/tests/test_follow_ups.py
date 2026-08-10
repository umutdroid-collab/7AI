"""Yöneticinin takip notları ve ziyaret raporu filtreleri/dışa aktarımı."""

import io
from datetime import date, timedelta

from openpyxl import load_workbook

from tests.test_checkin_photos import create_checkin, make_photo


def make_hospital(client, admin, name, city=None):
    return client.post("/hospitals", json={"name": name, "city": city}, headers=admin).json()


def test_follow_up_survives_and_is_ordered_by_due_date(client, admin):
    today = date.today()
    client.post(
        "/follow-ups",
        json={"note": "Sonraki toplantıda konuş", "remind_on": (today + timedelta(days=5)).isoformat()},
        headers=admin,
    )
    urgent = client.post(
        "/follow-ups",
        json={"note": "Acil: hastane şikayetçi", "remind_on": today.isoformat()},
        headers=admin,
    )
    assert urgent.status_code == 200, urgent.text

    items = client.get("/follow-ups", headers=admin).json()
    assert [i["note"] for i in items] == ["Acil: hastane şikayetçi", "Sonraki toplantıda konuş"]
    assert items[0]["days_until_due"] == 0
    assert items[1]["days_until_due"] == 5


def test_follow_up_carries_the_checkin_it_came_from(client, admin):
    checkin = create_checkin(client, admin).json()
    client.patch(f"/checkins/{checkin['id']}", json={"comment": "Ürünü beğenmediler"}, headers=admin)

    created = client.post(
        "/follow-ups",
        json={
            "note": "Beğenmeme sebebini sor",
            "checkin_id": checkin["id"],
            "remind_on": date.today().isoformat(),
        },
        headers=admin,
    ).json()

    assert created["checkin"]["id"] == checkin["id"]
    assert created["checkin"]["comment"] == "Ürünü beğenmediler"
    assert created["checkin"]["hospital_name"] == "H"


def test_completed_follow_ups_leave_the_list_but_are_not_deleted(client, admin):
    created = client.post(
        "/follow-ups", json={"note": "Takip", "remind_on": date.today().isoformat()}, headers=admin
    ).json()

    done = client.patch(f"/follow-ups/{created['id']}", json={"is_done": True}, headers=admin).json()
    assert done["is_done"] is True
    assert done["done_at"] is not None

    assert client.get("/follow-ups", headers=admin).json() == []
    assert len(client.get("/follow-ups", params={"include_done": True}, headers=admin).json()) == 1

    # Geri açılınca tamamlanma anı temizlenmeli, yoksa eski tarih yanlış bilgi verir.
    reopened = client.patch(f"/follow-ups/{created['id']}", json={"is_done": False}, headers=admin).json()
    assert reopened["done_at"] is None


def test_blank_note_is_rejected(client, admin):
    r = client.post("/follow-ups", json={"note": "   ", "remind_on": date.today().isoformat()}, headers=admin)
    assert r.status_code == 422


def test_follow_ups_are_admin_only(client, admin, employee):
    created = client.post(
        "/follow-ups", json={"note": "Takip", "remind_on": date.today().isoformat()}, headers=admin
    ).json()
    assert client.get("/follow-ups", headers=employee).status_code == 403
    assert client.post(
        "/follow-ups", json={"note": "X", "remind_on": date.today().isoformat()}, headers=employee
    ).status_code == 403
    assert client.patch(f"/follow-ups/{created['id']}", json={"is_done": True}, headers=employee).status_code == 403
    assert client.delete(f"/follow-ups/{created['id']}", headers=employee).status_code == 403


def test_due_follow_ups_appear_in_the_daily_digest(client, admin):
    """Notun amacı "unutmayayım"; hatırlatma günlük özette yapılmalı."""
    from app.database import SessionLocal
    from app.services.daily_digest import collect_digest

    today = date.today()
    client.post(
        "/follow-ups", json={"note": "Bugünkü not", "remind_on": today.isoformat()}, headers=admin
    )
    client.post(
        "/follow-ups",
        json={"note": "Gelecek hafta", "remind_on": (today + timedelta(days=7)).isoformat()},
        headers=admin,
    )

    db = SessionLocal()
    try:
        digest = collect_digest(db)
    finally:
        db.close()

    assert [f.note for f in digest["follow_ups"]] == ["Bugünkü not"]


def test_done_follow_ups_stay_out_of_the_digest(client, admin):
    from app.database import SessionLocal
    from app.services.daily_digest import collect_digest

    created = client.post(
        "/follow-ups", json={"note": "Bitti", "remind_on": date.today().isoformat()}, headers=admin
    ).json()
    client.patch(f"/follow-ups/{created['id']}", json={"is_done": True}, headers=admin)

    db = SessionLocal()
    try:
        assert collect_digest(db)["follow_ups"] == []
    finally:
        db.close()


# --- Ziyaret raporu: hastane filtresi ve Excel ---


def test_checkins_can_be_filtered_by_hospital(client, admin):
    a = make_hospital(client, admin, "A Hastanesi")
    b = make_hospital(client, admin, "B Hastanesi")
    for hospital in (a, b, b):
        client.post(
            "/checkins",
            data={"hospital_id": hospital["id"]},
            files={"photo": ("c.jpg", make_photo(400, 300), "image/jpeg")},
            headers=admin,
        )

    only_b = client.get("/checkins", params={"hospital_id": b["id"]}, headers=admin).json()
    assert len(only_b) == 2
    assert {c["hospital"]["name"] for c in only_b} == {"B Hastanesi"}


def test_checkin_export_respects_filters_and_omits_photos(client, admin):
    hospital = make_hospital(client, admin, "Rapor Hastanesi", city="İstanbul")
    client.post(
        "/checkins",
        data={"hospital_id": hospital["id"], "comment": "Ziyaret notu"},
        files={"photo": ("c.jpg", make_photo(400, 300), "image/jpeg")},
        headers=admin,
    )

    r = client.get("/checkins/export", params={"hospital_id": hospital["id"]}, headers=admin)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    ws = load_workbook(io.BytesIO(r.content)).active
    headers = [c.value for c in ws[1]]
    assert "Not" in headers and "Hastane" in headers
    assert not any("foto" in (h or "").lower() for h in headers)

    row = {headers[i]: c.value for i, c in enumerate(ws[2])}
    assert row["Hastane"] == "Rapor Hastanesi"
    assert row["Şehir"] == "İstanbul"
    assert row["Not"] == "Ziyaret notu"


def test_export_date_range_excludes_outside_visits(client, admin):
    hospital = make_hospital(client, admin, "Tarih Hastanesi")
    client.post(
        "/checkins",
        data={"hospital_id": hospital["id"]},
        files={"photo": ("c.jpg", make_photo(400, 300), "image/jpeg")},
        headers=admin,
    )

    yesterday = (date.today() - timedelta(days=1)).isoformat()
    r = client.get(
        "/checkins/export", params={"start_date": yesterday, "end_date": yesterday}, headers=admin
    )
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws.max_row == 1  # yalnızca başlık satırı

    today = date.today().isoformat()
    r = client.get("/checkins/export", params={"start_date": today, "end_date": today}, headers=admin)
    assert load_workbook(io.BytesIO(r.content)).active.max_row == 2


def test_employee_export_only_contains_their_own_visits(client, admin, employee):
    """Yetki filtresi listede olduğu gibi dışa aktarımda da geçerli olmalı."""
    hospital = make_hospital(client, admin, "Ortak Hastane")
    client.post(
        "/checkins",
        data={"hospital_id": hospital["id"]},
        files={"photo": ("c.jpg", make_photo(400, 300), "image/jpeg")},
        headers=admin,
    )

    r = client.get("/checkins/export", headers=employee)
    assert r.status_code == 200
    assert load_workbook(io.BytesIO(r.content)).active.max_row == 1  # admin'in ziyareti görünmemeli
