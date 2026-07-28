"""Fatura filtreleri, sıralama, silme ve Excel raporu."""

import io
from datetime import date, timedelta

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models import Invoice, InvoiceStatus, Notification, NotificationRead


def _seed(**kwargs):
    db = SessionLocal()
    inv = Invoice(file_path="", source_filename=kwargs.pop("source_filename", "x.pdf"), **kwargs)
    db.add(inv)
    db.commit()
    inv_id = inv.id
    db.close()
    return inv_id


def test_upcoming_sorted_soonest_first(client, admin):
    today = date.today()
    for n, days in [("U20", 20), ("U3", 3), ("U10", 10)]:
        _seed(invoice_number=n, due_date=today + timedelta(days=days))
    r = client.get("/invoices", params={"upcoming_only": True}, headers=admin)
    assert [i["invoice_number"] for i in r.json()] == ["U3", "U10", "U20"]


def test_overdue_sorted_most_overdue_first(client, admin):
    today = date.today()
    for n, days in [("O5", 5), ("O30", 30), ("O15", 15)]:
        _seed(invoice_number=n, due_date=today - timedelta(days=days))
    r = client.get("/invoices", params={"overdue_only": True}, headers=admin)
    assert [i["invoice_number"] for i in r.json()] == ["O30", "O15", "O5"]


def test_paid_invoices_excluded_from_due_tabs(client, admin):
    today = date.today()
    _seed(invoice_number="ODENDI", due_date=today - timedelta(days=5), status=InvoiceStatus.PAID)
    _seed(invoice_number="ACIK", due_date=today - timedelta(days=5))
    r = client.get("/invoices", params={"overdue_only": True}, headers=admin)
    assert [i["invoice_number"] for i in r.json()] == ["ACIK"]
    # ama Tümü listesinde görünmeye devam etmeli
    assert "ODENDI" in [i["invoice_number"] for i in client.get("/invoices", headers=admin).json()]


def test_default_list_sorted_by_invoice_number_desc(client, admin):
    for n in ["MEA001", "MEA003", "MEA002"]:
        _seed(invoice_number=n)
    r = client.get("/invoices", headers=admin)
    assert [i["invoice_number"] for i in r.json()] == ["MEA003", "MEA002", "MEA001"]


def test_delete_removes_related_notifications_and_reads(client, admin):
    inv_id = _seed(invoice_number="SIL")
    db = SessionLocal()
    n = Notification(invoice_id=inv_id, title="vade", message="m")
    db.add(n)
    db.flush()
    db.add(NotificationRead(notification_id=n.id, user_id=1))
    db.commit()
    db.close()

    assert client.delete(f"/invoices/{inv_id}", headers=admin).status_code == 204
    db = SessionLocal()
    assert db.query(Notification).count() == 0
    assert db.query(NotificationRead).count() == 0
    db.close()


def test_delete_is_admin_only(client, employee):
    inv_id = _seed(invoice_number="X")
    assert client.delete(f"/invoices/{inv_id}", headers=employee).status_code == 403


def test_excel_export_respects_filters(client, admin):
    today = date.today()
    _seed(invoice_number="GECMIS", due_date=today - timedelta(days=3), counterparty="ACME", amount=100.0)
    _seed(invoice_number="YAKLASAN", due_date=today + timedelta(days=3), counterparty="BETA", amount=200.0)

    def numbers(params):
        r = client.get("/invoices/export", params=params, headers=admin)
        assert r.status_code == 200
        ws = load_workbook(io.BytesIO(r.content)).active
        return [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]

    assert numbers({}) == ["YAKLASAN", "GECMIS"]
    assert numbers({"overdue_only": True}) == ["GECMIS"]
    assert numbers({"upcoming_only": True}) == ["YAKLASAN"]
    assert numbers({"q": "ACME"}) == ["GECMIS"]
